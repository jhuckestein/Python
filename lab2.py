import sys
import openpyxl


def OpenExcelFile(WBName, WSName):
	print("We are here to open", WBName, " Excel file and sheet ", WSName)
	try:
		WBHandle = openpyxl.load_workbook(WBName)
		if WSName in WBHandle.get_sheet_names():
			#We found the worksheet name in the workbook
			WSHandle = WBHandle.get_sheet_by_name(WSName)
			return(WBHandle, WSHandle)
		else:
			#WSName is not seen in the workbook
			return(WBHandle, None)
	except IOError:
		print("Was not able to open ", WBName)

def AddOutputTab(inputWB, OutputTabName):
	print("Create a new worksheet: ", OutputTabName, " in the workbook: ", inputWB)
	try:
		#If WSheet named OutputTabName exists delete it first
		# Check to see if the added tab name is in the list of sheets
		if OutputTabName in inputWB.get_sheet_names():
			#Have to create a tab handle to refer to object
			tabHandle = inputWB.get_sheet_by_name(OutputTabName)
			#Have to use the tabHandle to refer to object to remove
			inputWB.remove_sheet(tabHandle)
			print("Had to delete existing tab: ", OutputTabName)
		#Create new tab by name OutputTabName
		optab = inputWB.create_sheet(OutputTabName)
		return (optab)
	except IOError:
		print("Could not add new worksheet: ", OutputTabName)

def SummarizeByDeliveryType(dataWS, outputWS):
#write the output header first
	outputWS.cell(row=1, column=1).value = "Delivery Type"
	outputWS.column_dimensions['A'].width = 20
	outputWS.cell(row=1, column=2).value = "2017 DAYS"
	outputWS.column_dimensions['B'].width = 20
	outputWS.cell(row=2, column=1).value = "WEBEX"
	outputWS.cell(row=3, column=1).value = "In Person"
	#Iterate over the max number of rows in the worksheet to count column 5 based on value in column 6
	wcounter = 0
	icounter = 0
	for i in range(2, dataWS.max_row + 1):
		if(dataWS.cell(row=i, column=6).value == "Web-based"):
			wcounter += float(dataWS.cell(row=i, column=5).value)
		elif(dataWS.cell(row=i, column=6).value == "In-person"):
			icounter += float(dataWS.cell(row=i, column=5).value)
		else:
			print("Something other than Webex or ILT was scheduled")
	
	print("Number Webex= ", wcounter, " Number ILT= ", icounter)
	#Now write to the cells with the output calculated
	outputWS.cell(row=2, column=2).value = wcounter  #change the wcounter to str(wcounter) to make a string
	outputWS.cell(row=3, column=2).value = icounter  #same as above to make a string
	
	#Create a chart object, and define where the data comes from with headers
	chartObj = openpyxl.chart.PieChart() 
	
	cat = openpyxl.chart.Reference(outputWS, min_col=1, min_row=2, max_row=3)
	data = openpyxl.chart.Reference(outputWS, min_col=2, min_row=1, max_col=2, max_row=3)
	
	chartObj.add_data(data, titles_from_data=True)  #If the title is there use it
	chartObj.set_categories(cat)   #use the cat object above to set the categories
	chartObj.width = 10       #This is set in 10cm
	chartObj.height = 8       #Set in cm which is the size of the rectangle the chart appears in
	chartObj.dataLabels = openpyxl.chart.label.DataLabelList()  #These three lines set the chart to display percentages
	chartObj.dataLabels.showVal = True                          #this is set to False by default
	chartObj.dataLabels.showPercent = True                      #where it will still work but not display.
	
	chartObj.title = "Delivery Type Distribution"   #Title displayed on chart
	outputWS.add_chart(chartObj, 'C5')              #Anchors at C5 cell
	
	chart2 = openpyxl.chart.BarChart()
	chart2.add_data(data, titles_from_data=True)
	chart2.set_categories(cat)
	chart2.type = "col"
	chart2.style = 10
	chart2.y_axis.title = "Number of Days"
	chart2.x_axis.title = "Delivery Types"
	chart2.shape = 4
	outputWS.add_chart(chart2, "K5")
	
	
def SummarizeBySME(dataWS, outputWS):
#Print the header in the output worksheet
	outputWS.cell(row=1, column=1).value = "SME Name"
	outputWS.column_dimensions['A'].width = 20
	outputWS.cell(row=1, column=2).value = "Total Days"
	outputWS.column_dimensions['B'].width = 20
	outputWS.cell(row=1, column=3).value = "Web Days"
	outputWS.column_dimensions['C'].width = 20
	outputWS.cell(row=1, column=4).value = "In-person Days"
	outputWS.column_dimensions['D'].width = 20
	
	#First we need four empty lists SMEs, total days, Web Days, ILT Days to keep track of counts
	#Use option1 to create them
	SMEList = []
	TotalList = []  # index of SMEList will be used to store SME Total Days
	WebDaysList = [] # same index as above
	InpersonDaysList = [] # same index as above
	DTCol = 'F' #Delivery type column from main worksheet
	NumDaysCol = 'E' #Column number of days per delivery
	SMENameCol = 'N' #Column with SME name for primary instructor
	
	#We can use a for loop like SummarizeByDeliveryType to iterate through the master sheet
	for i in range(2, dataWS.max_row + 1):
		#process the data to map it to SME
		#See if SME name is already in the SMEList array, if not append to the list
		if (dataWS[SMENameCol+str(i)].value in SMEList):
			#Case where SME exists so find the index used for this SME and update
			idx = SMEList.index(dataWS[SMENameCol+str(i)].value) # Reference the position of SME in SMENameCol and get value of that position
			TotalList[idx] += float(dataWS[NumDaysCol+str(i)].value) #Reference the i position from main worksheet to calculated idx from previous line
			if (dataWS[DTCol+str(i)].value == "Web-based"):  # Using option1 to refer to cell in spreadsheet
				WebDaysList[idx] += float(dataWS[NumDaysCol+str(i)].value)
			elif (dataWS[DTCol+str(i)].value == "In-person"):
				InpersonDaysList[idx] += float(dataWS[NumDaysCol+str(i)].value)
			else:
				print("Something other than web-based or ILT was scheduled")
		#Now the SME is not in the list so append the SME to the end of the lists
		else:
			SMEList.append(dataWS[SMENameCol+str(i)].value)
			TotalList.append(float(dataWS[NumDaysCol+str(i)].value))
			if (dataWS[DTCol+str(i)].value == "Web-based"):  # Using option1 to refer to cell in spreadsheet
				WebDaysList.append(float(dataWS[NumDaysCol+str(i)].value))
				InpersonDaysList.append(0.0)  #Keep the indexes equal as the Inperson will have a hole if not initialized
			elif (dataWS[DTCol+str(i)].value == "In-person"):
				InpersonDaysList.append(float(dataWS[NumDaysCol+str(i)].value))
				WebDaysList.append(0.0)  #Keep the indexes equal as WebDays will have a hole if not initialized
			else:
				print("Something other than web-based or ILT was scheduled has not been appended")
		
	#Now we need to update the output sheet to display our results
	for i in range(0, len(SMEList)):
		outputWS.cell(row=i+2, column=1).value = str(SMEList[i])
		outputWS.cell(row=i+2, column=2).value = str(TotalList[i])
		outputWS.cell(row=i+2, column=3).value = str(WebDaysList[i])
		outputWS.cell(row=i+2, column=4).value = str(InpersonDaysList[i])

def SummarizeByCourse(dataWS, outputWS):
#Print the header in the output worksheet
	outputWS.cell(row=1, column=1).value = "Course Name"
	outputWS.column_dimensions['A'].width = 72
	outputWS.cell(row=1, column=2).value = "Total Days"
	outputWS.column_dimensions['B'].width = 20
	outputWS.cell(row=1, column=3).value = "Web Days"
	outputWS.column_dimensions['C'].width = 20
	outputWS.cell(row=1, column=4).value = "In-person Days"
	outputWS.column_dimensions['D'].width = 20	

	#First we need four empty lists Courses, total days, Web Days, ILT Days to keep track of counts
	#Use option1 to create them (WorksheetObj['A3'] = value)
	CourseList = []
	TotalList = []  # index of CourseList will be used to store Course Total Days
	WebDaysList = [] # same index as above
	InpersonDaysList = [] # same index as above
	DTCol = 'F' #Delivery type column from main worksheet
	NumDaysCol = 'E' #Column number of days per delivery
	CourseNameCol = 'C' #Column with Course names
	
	
	#We can use a for loop like SummarizeByDeliveryType to iterate through the master sheet
	for i in range(2, dataWS.max_row + 1):
		#process the data to map it to Course
		#See if Course name is already in the CourseList array, if not append to the list
		if (dataWS[CourseNameCol+str(i)].value in CourseList):
			#Case where Course exists so find the index used for this Course and update
			idx = CourseList.index(dataWS[CourseNameCol+str(i)].value) # Reference the position of Course in CourseNameCol and get value of that position
			TotalList[idx] += float(dataWS[NumDaysCol+str(i)].value) #Reference the i position from main worksheet to calculated idx from previous line
			if (dataWS[DTCol+str(i)].value == "Web-based"):  # Using option1 to refer to cell in spreadsheet
				WebDaysList[idx] += float(dataWS[NumDaysCol+str(i)].value)
			elif (dataWS[DTCol+str(i)].value == "In-person"):
				InpersonDaysList[idx] += float(dataWS[NumDaysCol+str(i)].value)
			else:
				print("Something other than web-based or ILT was scheduled")
		#Now the Course is not in the list so append the Course to the end of the lists
		else:
			CourseList.append(dataWS[CourseNameCol+str(i)].value)
			TotalList.append(float(dataWS[NumDaysCol+str(i)].value))
			if (dataWS[DTCol+str(i)].value == "Web-based"):  # Using option1 to refer to cell in spreadsheet
				WebDaysList.append(float(dataWS[NumDaysCol+str(i)].value))
				InpersonDaysList.append(0.0)  #Keep the indexes equal as the Inperson will have a hole if not initialized
			elif (dataWS[DTCol+str(i)].value == "In-person"):
				InpersonDaysList.append(float(dataWS[NumDaysCol+str(i)].value))
				WebDaysList.append(0.0)  #Keep the indexes equal as WebDays will have a hole if not initialized
			else:
				print("Something other than web-based or ILT was scheduled has not been appended")

	#Now we need to update the output sheet to display our results
	for i in range(0, len(CourseList)):
		outputWS.cell(row=i+2, column=1).value = str(CourseList[i])
		outputWS.cell(row=i+2, column=2).value = str(TotalList[i])
		outputWS.cell(row=i+2, column=3).value = str(WebDaysList[i])
		outputWS.cell(row=i+2, column=4).value = str(InpersonDaysList[i])

def SummarizeCourseInstance(dataWS, outputWS):
#Print the header in the output worksheet
	outputWS.cell(row=1, column=1).value = "Course Name"
	outputWS.column_dimensions['A'].width = 72
	outputWS.cell(row=1, column=2).value = "Duration"
	outputWS.column_dimensions['B'].width = 10
	outputWS.cell(row=1, column=3).value = "Total Days"
	outputWS.column_dimensions['C'].width = 12
	outputWS.cell(row=1, column=4).value = "Instances Recorded"
	outputWS.column_dimensions['D'].width = 20
	outputWS.cell(row=1, column=5).value = "Instances Taught"
	outputWS.column_dimensions['E'].width = 20

	#First we need empty lists Courses, total days, duration, instances recorded
	#Instances recorded is a count from the Master Scheduling Report, while Instances Taught
	#will be calculated after all lists have been iterated through and should match Instances Recorded.
	#Use option1 to create them (WorksheetObj['A3'] = value)
	CourseList = []
	DurationList = [] #index mapped to CourseList
	TotalList = []  # index of CourseList will be used to store Course Total Days
	InstRecorded = [] # same index as above
	InstTaught = [] # same index as above
	#DTCol = 'F' #Delivery type column from main worksheet
	NumDaysCol = 'E' #Column number of days per delivery
	CourseNameCol = 'C' #Column with Course names
	
	#Iterate through the Master Sheet with a for loop like other summaries
	for i in range(2, dataWS.max_row + 1):
		#process the data to map it to Course
		#See if Course name is already in the CourseList array, if not append to the list
		if (dataWS[CourseNameCol+str(i)].value in CourseList):
			#Case where Course exists so find the index used for this Course and update
			idx = CourseList.index(dataWS[CourseNameCol+str(i)].value) # Reference the position of Course in CourseNameCol and get value of that position
			TotalList[idx] += float(dataWS[NumDaysCol+str(i)].value) #Reference the i position from main worksheet to calculated idx from previous line
			InstRecorded[idx] += 1
		#Now the Course is not in the list so append the Course to the end of the lists
		#In this case we also initialize the DurationList.append(length of course) and InstRecorded.append(1 delivery)
		else:
			CourseList.append(dataWS[CourseNameCol+str(i)].value)
			TotalList.append(float(dataWS[NumDaysCol+str(i)].value))
			DurationList.append(float(dataWS[NumDaysCol+str(i)].value))  #Fill in the length of the course once so we don't overwrite a bunch of times above
			InstRecorded.append(1)   #This is the first time we found the course, so number of deliveries=1
	
	#At this point the first four arrays are populated, but as a check we need to iterate through the populated
	#arrays, and calculate InstTaught where (TotalList[i])/(DurationList[i]) should = InstRecorded[i]
	for i in range(0, len(CourseList)):
		InstTaught.append(TotalList[i]/DurationList[i])
		
	#Now we need to update the output sheet to display our results
	for i in range(0, len(CourseList)):
		outputWS.cell(row=i+2, column=1).value = str(CourseList[i])
		outputWS.cell(row=i+2, column=2).value = str(DurationList[i])
		outputWS.cell(row=i+2, column=3).value = str(TotalList[i])
		outputWS.cell(row=i+2, column=4).value = str(InstRecorded[i])
		outputWS.cell(row=i+2, column=5).value = str(InstTaught[i])

def SMEUniqueCourses(dataWS, outputWS):
#Print the header in the output worksheet
	outputWS.cell(row=1, column=1).value = "SME Name"
	outputWS.column_dimensions['A'].width = 20
	outputWS.cell(row=1, column=2).value = "Number of Courses"
	outputWS.column_dimensions['B'].width = 20
	outputWS.cell(row=1, column=3).value = "Unique Courses"
	outputWS.column_dimensions['C'].width = 72
	
	#Now we need to initialize a list of SMEs, their courses, and rownumber (for printing format)
	SMEList = []    #Each SME appears only once
	CourseList = [] #The index of each SME in the list is used, and a course string is built
	rownumber = 2   #Used for printing out the results on the worksheet below
	SMENameCol = 'N' #Column with SME name for primary instructor
	CourseNameCol = 'C' #Column with Course names
	SMECourses = []  #Used for processing an individual SMEs courses into a set
	SMECourseSet = [] #Used for processing with above
	SMECourseList = [] #Final processed non-redundant course list per SME
	
	
	#Now iterate through the master sheet and build the SMEList and CourseList
	for i in range(2, dataWS.max_row + 1):
		#process the data to map it to SME
		#See if SME name is already in the SMEList array, if not append to the list
		if (dataWS[SMENameCol+str(i)].value in SMEList):
			#Case where SME exists so find the index used for this SME
			idx = SMEList.index(dataWS[SMENameCol+str(i)].value) # Reference the position of SME in SMENameCol and get value of that position
			
			#Now retrieve the value of the current CourseList for the SME as a string
			courseString = str(CourseList[idx])
			#print("current courseString =", courseString)
			#print("from the master sheet ", dataWS[CourseNameCol+str(i)].value)
			
			#Next append the master worksheet entry to the end of the courseString
			courseString = ','.join([courseString, dataWS[CourseNameCol+str(i)].value])  
			#print("appended courseString =", courseString)
			
			#Replace the CourseList entry with the newly appended courseString string
			CourseList.pop(idx)
			CourseList.insert(idx, courseString)
			#print("current CourseList =", CourseList)
			
		#Now the SME is not in the list so append the SME to the end of the lists
		else:
			SMEList.append(dataWS[SMENameCol+str(i)].value)
			CourseList.append(str(dataWS[CourseNameCol+str(i)].value))
			
	
	#At this point we have two lists indexed to each other: SMEList which is a single
	#list of the SMEs, and CourseList which is a string of courses per/SME by commas.
	#Next iterate over the range of SMEs in the list to process and print each of
	#their unique courses onto the worksheet.
	
	for SME in range(0, len(SMEList)):
		#Retrieve the string of courses from CourseList and split it into a list
		#which will have redundancy.
		SMECourses.extend(CourseList[SME].split(','))
		
		#Make a set out of the list to remove redundancy
		SMECourseSet = set(SMECourses)
		
		#Sets don't allow indexing which we need for print formatting so turn the
		#set back into a list without redundancy.
		SMECourseList = [*SMECourseSet]
		
		#Now print out the results to the worksheet
		for i in range(0, len(SMECourseSet)):
			#If this is the first line then print the SME name and number of courses
			if (i == 0):
				outputWS.cell(row=rownumber, column=1).value = str(SMEList[SME])
				outputWS.cell(row=rownumber, column=2).value = str(len(SMECourseList))
				outputWS.cell(row=rownumber, column=3).value = str(SMECourseList[i])
				rownumber += 1
			#If this is not the first line then just print the unique course name
			else:
				outputWS.cell(row=rownumber, column=3).value = str(SMECourseList[i])
				rownumber += 1
		
		#Now clear out the SMECourses list and SMECourseSet to re-use for next SME
		SMECourses.clear()
		SMECourseSet.clear()
		SMECourseList.clear()
	
			


	
#main program starts here
###########################
#We are getting two returned values from OpenExcelFile, hence the two 
#variables in the print statement.
TMSWorkbook, MSRWorksheet = OpenExcelFile(sys.argv[1], "Master Scheduling Report")
print("Maximum rows =", MSRWorksheet.max_row, " Maximum columns =", MSRWorksheet.max_column)

#step1 - create the output tab or worksheet
OpTabh = AddOutputTab(TMSWorkbook, "By Delivery Type")

#step2 - create a summary table in the worksheet
SummarizeByDeliveryType(MSRWorksheet, OpTabh)


#step3 - create an output tab by SME Summary
OpTabh = AddOutputTab(TMSWorkbook, "By SME")
SummarizeBySME(MSRWorksheet, OpTabh)

#step4 - create an output tab by Course Summary (course, web, ILT)
OpTabh = AddOutputTab(TMSWorkbook, "By Course")
SummarizeByCourse(MSRWorksheet, OpTabh)

#step5 - create an output tab by Course Instances (course, total days, duration, instances recorded, instances taught)
OpTabh = AddOutputTab(TMSWorkbook, "Course Instances")
SummarizeCourseInstance(MSRWorksheet, OpTabh)

#step6 - create an output tab by SME listing their unique courses
OpTabh = AddOutputTab(TMSWorkbook, "SME Courses")
SMEUniqueCourses(MSRWorksheet, OpTabh)


TMSWorkbook.save(sys.argv[1])
print("successfully saved the excel workbook")

#######################################Left over notes
#print(TMSWorkbook.get_sheet_names())
# sheetobjectlist = TMSWorkbook.get_sheet_names()
# for s in sheetobjectlist:
	# print(s)