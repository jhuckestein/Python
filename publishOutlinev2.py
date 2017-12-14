###############################################
#JH - 12/2017
#Problem Description:
#Take a course plan in Excel format and extract the course outline
#from the corresponding worksheet tabs.  The information gathered will be used
#to populate a PDF Course Outline file with the appropriate formatting.
###############################################
#
#Key fields needed:
#		Course Name - 1 (text)
#		Course Number - 2 (text)
#		Duration -2 (text)
#		Course Type -2 (text)
#		Course Description - 3 (text) : Top Frame
#		Course Audience - 4 (text) : Left Frame
#		Learning Objectives - 5 (bulleted text) : Left Frame
#		Suggested Prerequisites - 6 (bulleted text) : Left Frame
#		Required Equipment - 7 (text) : Left Frame
#		Course Outline - 8 (bulleted text) : Center Frame (ch1-4), Right Frame (ch5-8)
###############################################
#
#Design outline:
#	1) Run from the command line with inputs for xlsm file or all files in the directory
#	2) Open a filehandler for the xlsm file, and use the v2.1 defined worksheets
#	3) Read the key fields needed
#	4) Using an ouput PDF template create a PDF Outline  (ReportLab, PyPDF2, Pdfrw) ReportLab was chosen.
#      - make a new text pdf to be used as a watermark, and format the text boxes.
#	   - watermark the new text pdf file onto the template pdf file
#	   - outputOutline is the new PDF finished product outline
################################################

#I'm just importing a bunch of options right now to prove a concept.
import sys
import openpyxl
import os
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, Frame
from PyPDF2 import PdfFileWriter, PdfFileReader

def OpenExcelFile(WBName, WSName, APIName):
	try:
		WBHandle = openpyxl.load_workbook(WBName)
		if WSName in WBHandle.get_sheet_names():
			#We found the worksheet name in the workbook
			WSHandle = WBHandle.get_sheet_by_name(WSName)
			if APIName in WBHandle.get_sheet_names():
				#We found the 0-ProjectAPI worksheet in the workbook
				APIHandle = WBHandle.get_sheet_by_name(APIName)
				return(WBHandle, WSHandle, APIHandle)
			else:
				print("Was not able to find ", APIName)
		else:
			print("Was not able to find ", WSName)
			sys.exit()
	except IOError:
		print("Was not able to open ", WBName)
		sys.exit()

def CollectTextStrings(OH, AH):
	cseName = str(AH['C6'].value)
	cseNumber = str(AH['C7'].value)
	durat = str(AH['C21'].value) + ' ' + AH['C22'].value
	cseType = str(AH['C23'].value)
	desc = str(OH['B23'].value)
	aud = str(OH['B32'].value)
	reqs = str(OH['B36'].value)
	return cseName, cseNumber, durat, cseType, desc, aud, reqs

#Note: GetObjectives is currently developed for 1-Outline as of 12/13/2017.  This means
#that the column and row numbers could change at some future point.		
def GetObjectives(WSHandle):
	objectives = []
	i = 40  #This value is used because 1-Outline currently has Learning Objectives beginning at C40-C47
	while WSHandle.cell(row = i, column = 3).value:
		objectives.append(WSHandle.cell(row = i, column = 3).value)
		i += 1
	#Now see if the list is populated, if not append a string to the first value
	if objectives:
		return objectives
	else:
		objectives.append(' ')
		return objectives

#Note: The current course outline populates 1-Outline on D101-D105.
def GetPreReqs(WSHandle):
	preReqs = []
	i = 101  #This value is used because 1-Outline currently has Prerequisites beginning at D101
	while WSHandle.cell(row = i, column = 4).value:
		preReqs.append(WSHandle.cell(row = i, column = 4).value)
		i += 1
	#Now see if the list is populated, if not append a string to the first value
	if preReqs:
		return preReqs
	else:
		preReqs.append(' ')
		return preReqs
	
		
#Note: GetOutline is currently developed for 1RS-Outline as of 12/7/2017.  This means
#that the column and row numbers could change at some future point.

def GetOutline(WSHandle):
	Coutline = []    #Center pane outline Ch1-4
	Routline = []    #Right pane outline Ch5-8
	CCont = []    #Keeps track of major headers for printing
	RCont = []    #Keeps track of major headers for printing
	lineNum = 49     #initialized to starting line for course outlines
	for i in range(1, 5):
		lineNum += 1
		if WSHandle.cell(row = lineNum, column = 3).value:    #Case where we hit major header
			CCont.append('Major')
			Coutline.append(str(i) + ' ' + str(WSHandle.cell(row = lineNum, column = 3).value))
		for k in range (1, 6):   #Case where we check for minor header
			lineNum += 1
			if WSHandle.cell(row = lineNum, column = 4).value:
				CCont.append('minor')
				Coutline.append(str(i) + '.' + str(k) + ' ' + str(WSHandle.cell(row = lineNum, column = 4).value))
	lineNum = 73     #reset the initialization just to be sure to process right half of outline
	for i in range(5, 9):
		lineNum += 1
		if WSHandle.cell(row = lineNum, column = 3).value:
			RCont.append('Major')
			Routline.append(str(i) + ' ' + str(WSHandle.cell(row = lineNum, column = 3).value))
		for k in range (1, 6):
			lineNum += 1
			if WSHandle.cell(row = lineNum, column = 4).value:
				RCont.append('minor')
				Routline.append(str(i) + '.' + str(k) + ' ' + str(WSHandle.cell(row = lineNum, column = 4).value))
	return Coutline, CCont, Routline, RCont
			

#The next function creates a PDF text file from collected information to be used as a watermark
def CreateTextFile(tFile, cName, cNum, dur, cType, desc, aud, objectives, prereqs, required, cOutline, cControl, rOutline, rControl):
#Create a new PDF (called a Canvas) with our collected Outline
	c = canvas.Canvas(tFile, pagesize=letter)
	
	##Now that the file is created draw in the top header section to include the course name, course type,
	##the duration, and the course number.  Note: these are not flowable fields just plain text strings.
	c.setFont("Helvetica-Bold", 21)
	c.drawString(40, 570, cName)
	c.setFont("Helvetica", 12)
	c.drawString(40, 548, cType)
	c.drawString(175, 548, dur)
	c.drawString(330, 548, cNum)
	#c.setFont("Helvetica", 10.5)
	#c.drawString(40, 520, description)
	
	##This next block is for the top frame used for the description
	styles = getSampleStyleSheet()
	styleN = styles['Normal']
	story = []      #Note: a 'story' is a ReportLabs term for 'flowables'
	story.append(Paragraph(desc, styleN))
	fTop = Frame(0.5*inch, 6.2*inch, 10*inch, 1.25*inch, showBoundary = 0)
	fTop.addFromList(story, c)
	
	##This next block is for the left frame used for audience, objectives, pre-requisites, and equipment
	styleH = styles['Heading2']
	styleh = styles['Heading5']
	stylen = styles['BodyText']
	styleB = styles['Bullet']
	storyL = []
	storyL.append(Paragraph('Intended Audience', styleH))
	storyL.append(Paragraph(aud, styleN))
	storyL.append(Paragraph('Learning Objectives', styleH))
	for obj in range(0, len(objectives)):
		storyL.append(Paragraph('  -' + objectives[obj], styleB))
	storyL.append(Paragraph('Suggested Prerequisites', styleH))
	for pre in range(0, len(prereqs)):
		storyL.append(Paragraph('  -' + prereqs[pre], styleB))
	storyL.append(Paragraph('Required Equipment', styleH))
	storyL.append(Paragraph(required, styleB))
	
	#Note: previously we thought prereqs would be a bulleted list and now it is just text
	#for req in range(0, len(required)):
	#	storyL.append(Paragraph('  -' + required[req], styleB))
	
	fLeft = Frame(0.5*inch, inch, 4.7*inch, 5*inch, showBoundary = 0)
	fLeft.addFromList(storyL, c)
	
	##This next block separates the outline content into center and right frame stories
	storyC = []
	storyR = []
	BHCount = 0                ##This variable keeps track of the number of Big Headers encountered in the outline list
	storyC.append(Paragraph('Course Outline', styleH))
	for line in range(0, len(cOutline)):
		if cControl[line] == 'Major':
			storyC.append(Paragraph(cOutline[line], styleh))   #Use bold format for Major Headers
		else:
			storyC.append(Paragraph(cOutline[line], styleN))   #Use normal format for minor headers
	for line in range(0, len(rOutline)):
		if rControl[line] == 'Major':
			storyR.append(Paragraph(rOutline[line], styleh))   #Use bold format for Major Headers
		else:
			storyR.append(Paragraph(rOutline[line], styleN))   #Use normal format for minor headers
	
	##This next block populates the center frame (Big Headers 1-4)
	fCent = Frame(5.4*inch, 0.5*inch, 2.5*inch, 5.5*inch, showBoundary = 0)
	fCent.addFromList(storyC, c)
	
	##This next block is for the right frame (Big Headers 5-8)
	fRight = Frame(8*inch, 0.5*inch, 2.5*inch, 5.15*inch, showBoundary = 0)
	fRight.addFromList(storyR, c)
	
	c.save()
	

#The next function takes the text-file PDF and uses it as a watermark to be placed upon the 
#graphics template, and produces the outputOutline.pdf
def CreateOutline(template, inFile, outFile):
	output = PdfFileWriter()
	ipdf = PdfFileReader(open(template, 'rb'))
	wpdf = PdfFileReader(open(inFile, 'rb'))
	watermark = wpdf.getPage(0)
	
	for i in xrange(ipdf.getNumPages()):
		page = ipdf.getPage(i)
		page.mergePage(watermark)
		output.addPage(page)
		
	with open(outFile, 'wb') as f:
		output.write(f)
	
################################################
# Main Program begins below

OutlineTab = '1-Outline'   #If the tabs on the Course Outline in Excel are re-named this will change.
APITab = '1-Budgeting'      #As above this tracks the worksheet tab in the Excel Course Outline.
graphicsTemplate = 'Example-NWV.pdf'  #This variable holds the pdf file used as a graphics template
#TextFile = This variable is the name of the text file which will be used as a watermark set below
#outputFile = This is the name of the final outline produced as a PDF set below

if (len(sys.argv) > 1):
	if sys.argv[1] == 'all':  #case where a batch process of a directory is desired
		inputFiles = []
		inputFiles = os.listdir(".")
		# for dirpath, dirs, files in os.walk("."):
			# inputFiles.append(files)
		print("inputFiles =", inputFiles)
		for file in range(0, len(inputFiles)):
			if  inputFiles[file].endswith('xlsm'):  #case where a file in the directory is the correct xlsm file
				##Set the watermark file and the finished outline to match the given filename
				currentFile = inputFiles[file]
				TextFile = 'text' + currentFile.replace("xlsm", "pdf")
				outputFile = currentFile.replace("xlsm", "pdf")
				
				##Attempt to open the workbook and sheets
				try:
					WorkbookH, OutlineWksheetH, ApiH = OpenExcelFile(currentFile, OutlineTab, APITab)
					
					##Grab all of the key fields listed in the design section
					courseName, courseNumber, duration, courseType, description, audience, requirements = CollectTextStrings(OutlineWksheetH, ApiH)
					learnObjectives = list(GetObjectives(OutlineWksheetH))
					prerequisites = list(GetPreReqs(OutlineWksheetH))
					courseOutline, CControl, rightOutline, RControl = GetOutline(OutlineWksheetH)
				 
				
					#At this point I have all of the information scraped from the Excel
					#outline, and am ready to create a new PDF to watermark with the PDF Template
					try:
						CreateTextFile(TextFile, courseName, courseNumber, duration, courseType, description, audience, learnObjectives, prerequisites, requirements, courseOutline, CControl, rightOutline, RControl)
					except:
						print("Was not able to create a textfile.pdf to watermark the graphics template.")
					
					#Now merge the example.pdf with the template to get output
					try:
						CreateOutline(graphicsTemplate, TextFile, outputFile)
					except:
						print("Was not able to assemble the final Outline.pdf from the watermark and template.")
						
					#Tell the user the outline is complete
					print('Done processing ', currentFile)
				
				except:
					print('Was not able to open the requested file =', sys.argv[1])
			#End of the if test on line 241 - the else is do nothing
		#End of for loop on 240
	else:	                        #case where the user input a specific filename
		##Set the watermark file and the finished outline to match the given filename
		enteredFile = sys.argv[1]
		TextFile = 'text' + enteredFile.replace("xlsm", "pdf")
		outputFile = enteredFile.replace("xlsm", "pdf")
		
		##Attempt to open the workbook and sheets
		try:
			WorkbookH, OutlineWksheetH, ApiH = OpenExcelFile(sys.argv[1], OutlineTab, APITab)
			
			##Grab all of the key fields listed in the design section
			courseName, courseNumber, duration, courseType, description, audience, requirements = CollectTextStrings(OutlineWksheetH, ApiH)
			learnObjectives = list(GetObjectives(OutlineWksheetH))
			prerequisites = list(GetPreReqs(OutlineWksheetH))
			courseOutline, CControl, rightOutline, RControl = GetOutline(OutlineWksheetH)
		 
		
			#At this point I have all of the information scraped from the Excel
			#outline, and am ready to create a new PDF to watermark with the PDF Template
			try:
				CreateTextFile(TextFile, courseName, courseNumber, duration, courseType, description, audience, learnObjectives, prerequisites, requirements, courseOutline, CControl, rightOutline, RControl)
			except:
				print("Was not able to create a textfile.pdf to watermark the graphics template.")
			
			#Now merge the example.pdf with the template to get output
			try:
				CreateOutline(graphicsTemplate, TextFile, outputFile)
			except:
				print("Was not able to assemble the final Outline.pdf from the watermark and template.")
				
			#Tell the user the outline is complete
			print('Done')
		
		except:
			print('Was not able to open the requested file =', sys.argv[1])


#If enough arguments were not specified give the user a hint
else:
	print('Input Arguments <CourseOutline.xlsm> or <all>')
	
	
	


